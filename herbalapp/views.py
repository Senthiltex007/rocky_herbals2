#Add Left + Add Right# herbalapp/views.py ====================================================== CLEAN, CONSISTENT VIEWS FOR Rocky Herbals (tree + app) Modern Tree (uses auto_id) - 
# single complete file ======================================================

from decimal import Decimal
from collections import deque
from django.db import transaction
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from herbalapp.models import DailyIncomeReport
from herbalapp.utils.invite import validate_and_consume_invite

import csv
import openpyxl
# -------------------------
# STATIC / SIMPLE PAGES
# -------------------------

from django.shortcuts import render

def join(request):
    """
    Join Us page
    """
    return render(request, "join.html", {"page_title": "Join Us"})


# Charts (openpyxl)
from openpyxl.chart import BarChart, LineChart, Reference

# ==========================
# Models
# ==========================
from .models import (
    Member,
    Product,
    Payment,
    CommissionRecord,
    BonusRecord,
    Order,
    IncomeRecord,
)

# Optional RankReward if present in your app
try:
    from .models import RankReward
except Exception:
    RankReward = None

# ==========================
# Forms
# ==========================
from .forms import MemberForm, ProductForm

# ==========================
# Services / business logic
# ==========================
try:
    from .services import add_billing_commission
except Exception:
    add_billing_commission = None


# ======================================================
# CHART HELPERS (no responses or external calls here)
# ======================================================
def add_bv_wallet_chart(ws, start_row):
    """
    Inserts a stacked bar chart comparing BV and Wallet values.
    Expects 5 rows starting from start_row:
        Row: [Label, Value]
    """

    # Create stacked bar chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "BV vs Wallet Comparison"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Values"

    # Data range (BV + Wallet summary rows)
    # Expected rows: [Label, Value]
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 4)
    cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + 4)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Make chart look cleaner
    chart.legend.position = "r"   # right side
    chart.height = 12             # visual size tuning
    chart.width = 20

    # Place chart in sheet
    ws.add_chart(chart, "P5")


def add_rank_progression_chart(ws, start_row, end_row):
    """
    Inserts a line chart showing monthly rank progression.
    Expects rows in the format:
        [MonthLabel, Count]
    """

    # Create line chart
    chart = LineChart()
    chart.title = "Monthly Rank Progression"
    chart.style = 13
    chart.y_axis.title = "Members"
    chart.x_axis.title = "Month"

    # Data range: rank counts per month
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Make chart visually clean
    chart.height = 12
    chart.width = 22
    chart.legend.position = "r"  # right side legend

    # Place chart in sheet
    ws.add_chart(chart, "P20")


def add_combo_chart(ws, bv_start_row, wallet_start_row, rank_start_row, rank_end_row):
    """
    Creates a combined chart:
    - Stacked Bar Chart for BV + Wallet summary
    - Line Chart for Rank Progression
    Expects rows in the format:
        [Label, Value]
    """

    # ============================
    # ✅ 1. Bar Chart (BV + Wallet)
    # ============================
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.title = "BV + Wallet Overview"
    bar.y_axis.title = "Values"
    bar.x_axis.title = "Category"

    # Combine BV + Wallet rows vertically
    bv_wallet_data = Reference(
        ws,
        min_col=2,
        min_row=bv_start_row,
        max_row=wallet_start_row + 1
    )
    bv_wallet_cats = Reference(
        ws,
        min_col=1,
        min_row=bv_start_row,
        max_row=wallet_start_row + 1
    )

    bar.add_data(bv_wallet_data, titles_from_data=False)
    bar.set_categories(bv_wallet_cats)

    # Make bar chart visually clean
    bar.height = 12
    bar.width = 22
    bar.legend.position = "r"


    # ============================
    # ✅ 2. Line Chart (Rank Trend)
    # ============================
    line = LineChart()
    line.title = "Rank Progression Trend"
    line.style = 13
    line.y_axis.title = "Members"
    line.x_axis.title = "Month"

    rank_data = Reference(
        ws,
        min_col=2,
        min_row=rank_start_row,
        max_row=rank_end_row
    )
    rank_cats = Reference(
        ws,
        min_col=1,
        min_row=rank_start_row,
        max_row=rank_end_row
    )

    line.add_data(rank_data, titles_from_data=False)
    line.set_categories(rank_cats)

    # Make line chart clean
    line.y_axis.majorGridlines = None


    # ============================
    # ✅ 3. Combine Charts
    # ============================
    bar += line


from openpyxl.chart import BarChart, LineChart, Reference
from django.db.models import QuerySet


def add_dashboard_sheet(wb, members):
    """
    Creates a Dashboard sheet with:
    - Overall summary
    - BV summary
    - Wallet summary
    - Rank summary
    - Rank Distribution Chart
    - BV vs Wallet Chart
    """

    # ==================================================
    # ✅ SAFETY: members can be queryset OR list
    # ==================================================
    if isinstance(members, QuerySet):
        members_qs = members
        members_list = list(members)
    else:
        members_list = list(members)
        members_qs = Member.objects.filter(
            id__in=[m.id for m in members_list]
        )

    ws = wb.create_sheet(title="Dashboard")

    # ==================================================
    # ✅ OVERALL SUMMARY
    # ==================================================
    ws.append(["Dashboard Summary"])
    ws.append(["Total Members", len(members_list)])

    total_left_bv = sum(
        (m.get_bv_counts()[0] or 0) for m in members_list
    )
    total_right_bv = sum(
        (m.get_bv_counts()[1] or 0) for m in members_list
    )

    ws.append(["Total Left BV", total_left_bv])
    ws.append(["Total Right BV", total_right_bv])
    ws.append(["Total Matched BV", min(total_left_bv, total_right_bv)])

    total_repurchase = sum(
        getattr(m, "repurchase_wallet", 0) or 0
        for m in members_list
    )
    total_flash = sum(
        getattr(m, "flash_wallet", 0) or 0
        for m in members_list
    )

    ws.append(["Total Repurchase Wallet", total_repurchase])
    ws.append(["Total Flash Wallet", total_flash])

    # ==================================================
    # ✅ RANK SUMMARY (DYNAMIC – NO HARD CODING)
    # ==================================================
    ws.append([])
    ws.append(["Rank Summary"])

    rank_start_row = ws.max_row + 1

    ranks = (
        members_qs
        .values_list("current_rank", flat=True)
        .exclude(current_rank__isnull=True)
        .exclude(current_rank="")
        .distinct()
    )

    if not ranks:
        ws.append(["No Rank Data", 0])
    else:
        for rank in ranks:
            count = members_qs.filter(current_rank=rank).count()
            ws.append([rank, count])

    rank_end_row = ws.max_row

    # ==================================================
    # ✅ RANK DISTRIBUTION CHART
    # ==================================================
    chart1 = BarChart()
    chart1.title = "Rank Distribution"
    chart1.x_axis.title = "Rank"
    chart1.y_axis.title = "Members"
    chart1.height = 12
    chart1.width = 20
    chart1.legend.position = "r"

    data = Reference(
        ws,
        min_col=2,
        min_row=rank_start_row,
        max_row=rank_end_row
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=rank_start_row,
        max_row=rank_end_row
    )

    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "E5")

    # ==================================================
    # ✅ BV vs WALLET CHART (NO HARD-CODED ROWS)
    # ==================================================
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"  # ✅ correct for single-series
    chart2.title = "BV vs Wallet Comparison"
    chart2.height = 12
    chart2.width = 20
    chart2.legend.position = "r"

    bv_wallet_start_row = 2
    bv_wallet_end_row = 1 + 1 + 5  # summary rows count (safe)

    data2 = Reference(
        ws,
        min_col=2,
        min_row=bv_wallet_start_row,
        max_row=bv_wallet_end_row
    )
    cats2 = Reference(
        ws,
        min_col=1,
        min_row=bv_wallet_start_row,
        max_row=bv_wallet_end_row
    )

    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "E20")

    # ==================================================
    # ✅ DONE – CLEAN & SAFE
    # ==================================================

# -------------------------
# HOME / STATIC PAGES
# -------------------------

from django.shortcuts import render


def home(request):
    """
    Home page
    """
    return render(request, "home.html")


def about(request):
    """
    About Us page
    """
    context = {
        "page_title": "About Us",
        "company": "Rocky Herbals",
        "description": (
            "We are a herbal products company focused on natural wellness "
            "and sustainable herbal solutions."
        ),
    }
    return render(request, "about.html", context)


def contact(request):
    """
    Contact Us page
    """
    context = {
        "page_title": "Contact Us",
        "email": "info@rockyherbals.com",
        "phone": "+91 81221 05779",
        "address": "Nambiyur, Tamil Nadu, India",
    }
    return render(request, "contact.html", context)


# -------------------------
# AUTH / MEMBER LOGIN
# -------------------------

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login


def member_login(request):
    """
    Member login view
    """

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        # ✅ Basic validation
        if not username or not password:
            messages.error(request, "Username and password are required.")
            return render(request, "member_login.html")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            # ✅ Optional: block inactive users
            if not user.is_active:
                messages.error(request, "Your account is inactive. Please contact admin.")
                return render(request, "member_login.html")

            login(request, user)
            return redirect("member_list")

        # ❌ Invalid credentials
        messages.error(request, "Invalid username or password.")

    return render(request, "member_login.html")


# -------------------------
# ADMIN LOGIN / DASHBOARD
# -------------------------

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum

from .models import Member, Payment, Product


# =========================
# 🔐 ADMIN LOGIN
# =========================
def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        # ✅ Basic validation
        if not username or not password:
            messages.error(request, "Username and password are required.")
            return render(request, "admin_login.html")

        user = authenticate(request, username=username, password=password)

        # ✅ Only staff users allowed
        if user and user.is_staff:
            if not user.is_active:
                messages.error(request, "Admin account is inactive.")
                return render(request, "admin_login.html")

            login(request, user)
            return redirect("admin_dashboard")

        messages.error(request, "Invalid admin credentials.")

    return render(request, "admin_login.html")


# =========================
# 🔒 ADMIN CHECK
# =========================
def is_admin(user):
    return user.is_authenticated and user.is_staff


# =========================
# 📊 ADMIN DASHBOARD
# =========================
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):

    # 👥 Members
    total_members = Member.objects.count()

    # 💳 Payments
    paid_members = Payment.objects.filter(status="Paid").count()
    unpaid_members = total_members - paid_members

    # 💰 Total income (null-safe)
    total_income = (
        Payment.objects
        .filter(status="Paid")
        .aggregate(total=Sum("amount"))
        .get("total") or 0
    )

    context = {
        "company_name": "Rocky Herbals",
        "total_members": total_members,
        "paid_members": paid_members,
        "unpaid_members": unpaid_members,
        "income": total_income,
        "products_count": Product.objects.count(),
        "income_count": DailyIncomeReport.objects.count(),
        "incomes": DailyIncomeReport.objects.all(),
    }

    return render(request, "admin_dashboard.html", context)

# -------------------------
# MEMBER CRUD / LIST
# -------------------------

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST

from .models import Member


# =========================
# 🔒 ADMIN CHECK
# =========================
def is_admin(user):
    return user.is_authenticated and user.is_staff


# =========================
# 👥 MEMBER LIST
# =========================
@login_required
@user_passes_test(is_admin)
def member_list(request):
    members = Member.objects.all().order_by("id")
    return render(request, "member_list.html", {"members": members})


# =========================
# ❌ DELETE MEMBER (ADMIN ONLY)
# =========================
@login_required
@user_passes_test(is_admin)
@require_POST
def delete_member(request, auto_id):

    # ✅ Support both numeric PK & auto_id
    if str(auto_id).isdigit():
        member = get_object_or_404(Member, id=auto_id)
    else:
        member = get_object_or_404(Member, auto_id=auto_id)

    # 🚫 Protect ROOT / Dummy members
    if member.auto_id in ["rocky001", "rocky001"]:
        messages.error(request, "❌ Root member cannot be deleted.")
        return redirect("member_list")

    # 🚫 Block if member has children
    if member.left_child or member.right_child():
        messages.error(
            request,
            f"❌ Cannot delete {member.name}. Remove left/right children first."
        )
        return redirect("member_list")

    # ✅ Detach safely from parent
    parent = member.parent
    if parent:
        if parent.left_child_id == member.id:
            parent.left_child = None
        elif parent.right_child_id == member.id:
            parent.right_child = None
        parent.save(update_fields=["left_child", "right_child"])

    member.delete()

    messages.success(request, f"✅ {member.name} deleted successfully!")
    return redirect("member_list")


# =========================
# 🔁 REPLACE / MOVE MEMBER
# =========================
@login_required
@user_passes_test(is_admin)
def replace_member(request, auto_id):

    # ✅ auto_id safe lookup
    member = get_object_or_404(Member, auto_id=auto_id)

    # 🚫 Protect root / dummy members
    if member.auto_id in ["rocky001", "rocky001"]:
        messages.error(request, "❌ Root member cannot be moved.")
        return redirect("member_list")

    # =========================
    # 🔁 POST → MOVE
    # =========================
    if request.method == "POST":
        new_parent_id = request.POST.get("new_parent")
        new_side = request.POST.get("side")

        if new_side not in ["L", "R"]:
            messages.error(request, "❌ Invalid side selected.")
            return redirect("replace_member", auto_id=member.auto_id)

        new_parent = get_object_or_404(Member, id=new_parent_id)

        # 🚫 Cannot move under itself
        if new_parent.id == member.id:
            messages.error(request, "❌ A member cannot be their own parent.")
            return redirect("replace_member", auto_id=member.auto_id)

        # 🚫 Prevent loop (downline check)
        def is_downline(root, target):
            if not root:
                return False
            if root.id == target.id:
                return True
            return (
                is_downline(root.left_child, target) or
                is_downline(root.right_child, target)
            )

        if is_downline(member, new_parent):
            messages.error(request, "❌ Cannot move under own downline.")
            return redirect("replace_member", auto_id=member.auto_id)

        # 🚫 Side occupied check
        if new_side == "L" and new_parent.left_child:
            messages.error(request, "❌ Left side already occupied.")
            return redirect("replace_member", auto_id=member.auto_id)

        if new_side == "R" and new_parent.right_child:
            messages.error(request, "❌ Right side already occupied.")
            return redirect("replace_member", auto_id=member.auto_id)

        # =========================
        # 🔐 ATOMIC TREE MOVE
        # =========================
        with transaction.atomic():

            # Detach from old parent
            old_parent = member.parent
            if old_parent:
                if old_parent.left_child_id == member.id:
                    old_parent.left_child = None
                elif old_parent.right_child_id == member.id:
                    old_parent.right_child = None
                old_parent.save(update_fields=["left_child", "right_child"])

            # Attach to new parent
            member.parent = new_parent
            member.side = new_side

            if new_side == "L":
                new_parent.left_child = member
            else:
                new_parent.right_child = member

            new_parent.save(update_fields=["left_child", "right_child"])
            member.save(update_fields=["parent", "side"])

        messages.success(request, "✅ Member moved successfully.")
        return redirect("member_list")

    # =========================
    # GET → SHOW FORM
    # =========================
    members = Member.objects.exclude(id=member.id)

    return render(
        request,
        "replace_member.html",
        {
            "member": member,
            "members": members,
        }
    )


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages


# -------------------------
# PRODUCT LIST
# -------------------------
@login_required
def product_list(request):
    products = Product.objects.all().order_by("-id")
    return render(request, "products.html", {"products": products})


# -------------------------
# ADD PRODUCT
# -------------------------
@login_required
def add_product(request):

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)

        # ✅ Optional: prevent duplicate product names
        name = request.POST.get("name")
        if Product.objects.filter(name__iexact=name).exists():
            messages.error(request, "❌ Product with this name already exists.")
            return render(request, "add_product.html", {"form": form})

        if form.is_valid():
            form.save()
            messages.success(request, "✅ Product added successfully!")
            return redirect("products")

        messages.error(request, "❌ Please correct the errors below.")

    else:
        form = ProductForm()

    return render(request, "add_product.html", {"form": form})


# -------------------------
# EDIT PRODUCT
# -------------------------
@login_required
def edit_product(request, product_id):

    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)

        # ✅ Optional: prevent duplicate names on edit
        name = request.POST.get("name")
        if Product.objects.filter(name__iexact=name).exclude(id=product.id).exists():
            messages.error(request, "❌ Another product with this name already exists.")
            return render(request, "add_product.html", {"form": form, "edit": True})

        if form.is_valid():
            form.save()
            messages.success(request, "✅ Product updated successfully!")
            return redirect("products")

        messages.error(request, "❌ Please correct the errors below.")

    else:
        form = ProductForm(instance=product)

    return render(request, "add_product.html", {"form": form, "edit": True})


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse


# -------------------------
# DELETE PRODUCT (SAFE)
# -------------------------
@login_required
def delete_product(request, product_id):

    product = get_object_or_404(Product, id=product_id)

    # ✅ Prevent deleting products linked to orders
    if Order.objects.filter(product=product).exists():
        messages.error(request, "❌ Cannot delete this product. It is linked to existing orders.")
        return redirect("products")

    product.delete()
    messages.success(request, "✅ Product deleted successfully!")
    return redirect("products")


# -------------------------
# CART VIEW
# -------------------------
@login_required
def cart_view(request):
    # Placeholder cart structure
    cart = request.session.get("cart", {})
    return render(request, "cart.html", {"cart": cart})


# -------------------------
# ADD TO CART
# -------------------------
@login_required
def add_to_cart(request, product_id):

    product = get_object_or_404(Product, id=product_id)

    # ✅ Initialize cart if not present
    cart = request.session.get("cart", {})

    # ✅ Add or increment quantity
    cart[str(product_id)] = cart.get(str(product_id), 0) + 1

    request.session["cart"] = cart
    messages.success(request, f"✅ {product.name} added to cart!")

    return redirect("cart")


# -------------------------
# REMOVE FROM CART
# -------------------------
@login_required
def remove_from_cart(request, product_id):

    cart = request.session.get("cart", {})

    if str(product_id) in cart:
        del cart[str(product_id)]
        request.session["cart"] = cart
        messages.success(request, "✅ Item removed from cart!")
    else:
        messages.error(request, "❌ Item not found in cart.")

    return redirect("cart")


# -------------------------
# CHECKOUT (PLACEHOLDER)
# -------------------------
@login_required
def checkout(request):
    return HttpResponse("Checkout Page - To be implemented")


from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .models import Payment, Commission, Member


# -------------------------
# MANUAL COMMISSION CREDIT (simple)
# -------------------------
@login_required
def credit_commission(request, auto_id):

    # Support both numeric PK and business auto_id
    if str(auto_id).isdigit():
        member = get_object_or_404(Member, id=auto_id)
    else:
        member = get_object_or_404(Member, auto_id=auto_id)

    messages.success(request, f"✅ Commission credited for {member.name}")
    return redirect("member_list")


# =====================================================
# PAYMENT APPROVE & STOCK COMMISSION GENERATOR
# =====================================================
@login_required
def approve_payment(request, payment_id):

    payment = get_object_or_404(Payment, id=payment_id)
    member = payment.member

    # ✅ Prevent double approval
    if payment.status == "approved":
        messages.warning(request, "⚠️ Payment already approved.")
        return redirect("payments_list")

    # ✅ Commission percentage mapping
    LEVEL_PERCENT = {
        "district": 7,
        "taluk": 5,
        "pincode": 3,
    }

    percentage = LEVEL_PERCENT.get(member.level, 0)
    commission_amount = (payment.amount * percentage) / 100

    # ✅ Create commission record only if eligible
    if percentage > 0:
        Commission.objects.create(
            member=member,
            payment=payment,
            commission_type=member.level,
            percentage=percentage,
            commission_amount=commission_amount
        )

    # ✅ Mark payment approved
    payment.status = "approved"
    payment.save()

    messages.success(request, f"✅ Payment approved and commission credited to {member.name}")
    return redirect("payments_list")


from django.urls import reverse

# ======================================================
# ROBUST TREE FUNCTIONS
# ======================================================
def _get_children(member):
    if not member:
        return None, None
    return member.left_child, member.right_child()


def build_tree_html(member):
    if not member:
        return ""

    left, right = _get_children(member)

    # ✅ Safe escaping
    safe_name = (member.name or "").replace("'", "\\'").replace('"', '\\"')
    safe_phone = (member.phone or "").replace("'", "\\'").replace('"', '\\"')
    safe_auto = str(member.auto_id)

    # ✅ Node HTML
    node_html = (
        f"<li>"
        f"<div class='member-box' "
        f"onclick=\"showMemberDetail({{id:'{member.auto_id}', name:'{safe_name}', phone:'{safe_phone}'}})\">"
        f"{safe_name} <br><small>ID: {safe_auto}</small>"
        f"</div>"
        f"<ul>"
    )

    # -------------------------
    # LEFT CHILD
    # -------------------------
    if left:
        node_html += build_tree_html(left)
    else:
        add_url = reverse('add_member_form') + f"?parent={member.auto_id}&side=left"
        node_html += (
            "<li><div class='member-box text-muted'>➕ Left<br>"
            f"<a href='{add_url}'>Add</a></div></li>"
        )

    # -------------------------
    # RIGHT CHILD
    # -------------------------
    if right:
        node_html += build_tree_html(right)
    else:
        add_url = reverse('add_member_form') + f"?parent={member.auto_id}&side=right"
        node_html += (
            "<li><div class='member-box text-muted'>➕ Right<br>"
            f"<a href='{add_url}'>Add</a></div></li>"
        )

    node_html += "</ul></li>"
    return node_html

from django.shortcuts import render, get_object_or_404
from herbalapp.models import Member

def tree_view(request, auto_id):
    # Normalize numeric ID → auto_id
    if str(auto_id).isdigit():
        try:
            pk_member = Member.objects.get(id=int(auto_id))
            auto_id = pk_member.auto_id
        except Member.DoesNotExist:
            return render(request, "tree_not_found.html", {"auto_id": auto_id})

    # Load member safely
    member = get_object_or_404(Member, auto_id=auto_id)

    # Corrected counts
    left_count = member.children.filter(side="left").count()
    right_count = member.children.filter(side="right").count()
    pairs = min(left_count, right_count)

    context = {
        "root_member": member,
        "left_count": left_count,
        "right_count": right_count,
        "pairs": pairs,
        "binary_income": getattr(member, "binary_income", 0),
        "sponsor_income": getattr(member, "sponsor_income", 0),
        "flashout": getattr(member, "flashout", 0),
        "binary_eligible": getattr(member, "binary_eligible", 0),
    }
    return render(request, "herbalapp/dynamic_tree.html", context)


# ======================================================
# PYRAMID / MLM TREE VIEW
# ======================================================
from django.shortcuts import render, get_object_or_404
from herbalapp.models import Member

def pyramid_view(request, auto_id):
    """
    Shows a pyramid/tree starting from a given member.
    Handles numeric IDs, missing members, and passes the root to template.
    """

    # Normalize numeric ID → auto_id
    if str(auto_id).isdigit():
        if str(auto_id) == "1":
            auto_id = "rocky001"  # Root member
        else:
            try:
                pk_member = Member.objects.get(id=int(auto_id))
                auto_id = pk_member.auto_id
            except Member.DoesNotExist:
                return render(request, "tree_not_found.html", {"auto_id": auto_id})

    # Load member safely
    root = Member.objects.filter(auto_id=auto_id).first()
    if not root:
        return render(request, "tree_not_found.html", {"auto_id": auto_id})

    # ✅ Prefetch children instead of left_child/right_child
    root = Member.objects.prefetch_related("children").get(auto_id=auto_id)

    # Render pyramid/tree HTML
    return render(request, "herbalapp/dynamic_tree.html", {
        "root_member": root
    })

from django.shortcuts import render, get_object_or_404
from django.urls import reverse
from .models import Member

from django.shortcuts import render, get_object_or_404
from django.db.models import Prefetch
from .models import Member

# ======================================================
# TREE VIEW HELPER
# ======================================================
def _normalize_member_id(auto_id):
    """Convert numeric ID → business auto_id."""
    if str(auto_id).isdigit():
        if str(auto_id) == "1":
            return "rocky001"  # Root member
        try:
            pk_member = Member.objects.get(id=int(auto_id))
            return pk_member.auto_id
        except Member.DoesNotExist:
            return None
    return auto_id


def member_not_found(request, member_id=None):
    return render(request, 'herbalapp/member_not_found.html', {'member_id': member_id})


# ======================================================
# MEMBER TREE VIEW
# ======================================================
def member_tree(request, auto_id):
    """Show tree starting from any member with prefetch optimization."""

    auto_id = _normalize_member_id(auto_id)
    if not auto_id:
        return render(request, "tree_not_found.html", {"auto_id": auto_id})

    root = get_object_or_404(
        Member.objects.prefetch_related(Prefetch("children")),
        auto_id=auto_id
    )

    return render(request, "herbalapp/dynamic_tree.html", {
        "root_member": root
    })


# ======================================================
# DYNAMIC TREE VIEW
# ======================================================
def dynamic_tree(request, auto_id):
    """Avatar-based dynamic tree with optimized prefetch."""

    auto_id = _normalize_member_id(auto_id)
    if not auto_id:
        return render(request, "tree_not_found.html", {"auto_id": auto_id})

    member = get_object_or_404(
        Member.objects.prefetch_related(Prefetch("children")),
        auto_id=auto_id
    )

    return render(request, "herbalapp/dynamic_tree.html", {
        "root_member": member
    })

from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from .models import Member

def member_detail_json(request, auto_id):
    # -------------------------
    # Normalize numeric → business auto_id
    # -------------------------
    if str(auto_id).isdigit():
        if str(auto_id) == "1":
            auto_id = "rocky001"
        else:
            try:
                pk_member = Member.objects.get(id=int(auto_id))
                auto_id = pk_member.auto_id
            except Member.DoesNotExist:
                return JsonResponse({"error": "Member not found"}, status=404)

    # -------------------------
    # Load member
    # -------------------------
    try:
        member = Member.objects.get(auto_id=auto_id)
    except Member.DoesNotExist:
        return JsonResponse({"error": "Member not found"}, status=404)

    # -------------------------
    # BV calculation
    # -------------------------
    bv_data = member.calculate_bv()  # Should return self_bv, left_bv, right_bv, total_bv

    # -------------------------
    # Last month BV
    # -------------------------
    last_month_start = (timezone.now().replace(day=1) - timedelta(days=1)).replace(day=1)
    last_month_end = (last_month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    last_month_orders = member.order_set.filter(
        status="Paid",
        created_at__date__gte=last_month_start,
        created_at__date__lte=last_month_end
    )
    last_month_bv = sum(o.product.bv_value * o.quantity for o in last_month_orders) if last_month_orders else 0

    # -------------------------
    # Sponsor / Binary / Flashout / Carry forward
    # -------------------------
    # These fields assume your Member model has helper methods to calculate per rules:
    # - member.get_sponsor_income()
    # - member.get_binary_income()
    # - member.get_flashout_units()
    # - member.get_carry_forward_pairs()

    sponsor_income = getattr(member, "get_sponsor_income", lambda: 0)()
    binary_income = getattr(member, "get_binary_income", lambda: 0)()
    flashout_units = getattr(member, "get_flashout_units", lambda: 0)()
    carry_forward_pairs = getattr(member, "get_carry_forward_pairs", lambda: 0)()

    # -------------------------
    # JSON response
    # -------------------------
    data = {
        "auto_id": member.auto_id,
        "name": member.name,
        "phone": member.phone,
        "self_bv": float(bv_data.get("self_bv", 0)),
        "left_bv": float(bv_data.get("left_bv", 0)),
        "right_bv": float(bv_data.get("right_bv", 0)),
        "total_bv": float(bv_data.get("total_bv", 0)),
        "last_month_bv": float(last_month_bv),
        "sponsor_income": float(sponsor_income),
        "binary_income": float(binary_income),
        "flashout_units": int(flashout_units),
        "carry_forward_pairs": int(carry_forward_pairs),
    }

    return JsonResponse(data)


from django.db.models import Prefetch
from .models import Member as _Member

# -------------------------
# Modern Tree (FINAL AVATAR VERSION)
# -------------------------

def _normalize_member_id(auto_id):
    """Convert numeric PK → business auto_id."""
    if str(auto_id).isdigit():
        if str(auto_id) == "1":
            return "rocky001"
        try:
            pk_member = _Member.objects.get(id=int(auto_id))
            return pk_member.auto_id
        except _Member.DoesNotExist:
            return None
    return auto_id


def member_tree_modern(request, auto_id):
    auto_id = _normalize_member_id(auto_id)
    if not auto_id:
        return _render(request, "tree_not_found.html", {"auto_id": auto_id})

    # ✅ Prefetch children for recursion
    root = _get(
        _Member.objects.prefetch_related(Prefetch("children")),
        auto_id=auto_id
    )

    return _render(request, "herbalapp/dynamic_tree.html", {
        "root_member": root
    })


from django.shortcuts import render as _render
from django.db.models import Prefetch
from .models import Member as _Member

def member_tree_modern_root(request):
    """
    Show the genealogy tree starting from rocky002 as root.
    """

    root = _Member.objects.filter(auto_id="rocky002").prefetch_related(
        Prefetch("children")
    ).first()

    if not root:
        return _render(request, "tree_not_found.html", {"message": "No root member found."})

    return _render(request, "herbalapp/dynamic_tree.html", {
        "root_member": root
    })


# -------------------------
# Edit Member (FINAL FIXED VERSION)
# -------------------------

from django.shortcuts import render as _render, get_object_or_404 as _get, redirect as _redirect
from .models import Member as _Member

def edit_member(request, auto_id):

    # ✅ Normalize numeric → business ID
    auto_id = _normalize_member_id(auto_id)
    if not auto_id:
        return _render(request, "tree_not_found.html", {"auto_id": auto_id})

    member = _get(_Member, auto_id=auto_id)

    if request.method == 'POST':
        member.name = request.POST.get('name')
        member.phone = request.POST.get('phone')
        member.aadhar = request.POST.get('aadhar')
        member.save()

        # ✅ Redirect to modern avatar tree
        return _redirect('member_tree_modern', auto_id=member.auto_id)

    return _render(request, 'edit_member.html', {'member': member})


from django.shortcuts import render as ___render, get_object_or_404 as ___get, redirect as ___redirect
from .models import Member as ___Member

def edit_sponsor(request, auto_id):

    # ✅ Normalize numeric → business auto_id
    if str(auto_id).isdigit():
        if str(auto_id) == "1":
            auto_id = "rocky001"
        else:
            try:
                pk_member = ___Member.objects.get(id=int(auto_id))
                auto_id = pk_member.auto_id
            except:
                return ___render(request, "tree_not_found.html", {"auto_id": auto_id})

    # ✅ Load member using business auto_id
    member = ___get(___Member, auto_id=auto_id)

    if request.method == 'POST':
        sponsor_auto_id = request.POST.get('sponsor_auto_id')

        # ✅ Load sponsor using auto_id
        sponsor = ___Member.objects.filter(auto_id=sponsor_auto_id).first()

        if sponsor:
            member.parent = sponsor
            member.save()

            # ✅ Redirect to modern avatar tree (CORRECT)
            return ___redirect('member_tree_modern', auto_id=member.auto_id)

    return ___render(request, 'edit_sponsor.html', {'member': member})


# ======================================================
# INCOME PAGE WITH FILTERS (FINAL VERSION) ✅ FIXED
# - Supports Date filter (DD-MM-YYYY + YYYY-MM-DD)
# - Supports min/max filter
# - Sends template context: run_date, reports, totals ✅
# ======================================================

from decimal import Decimal
from datetime import datetime
from django.shortcuts import render
from django.utils.timezone import localdate
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce

from herbalapp.models import DailyIncomeReport


def income_view(request):
    # -----------------------------
    # 1) Date filter
    # supports:
    # - UI input date picker: YYYY-MM-DD
    # - manual typed: DD-MM-YYYY
    # -----------------------------
    date_str = request.GET.get("date")
    if date_str:
        try:
            run_date = datetime.strptime(date_str, "%d-%m-%Y").date()
        except ValueError:
            try:
                run_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                run_date = localdate()
    else:
        run_date = localdate()

    # -----------------------------
    # 2) Base queryset by date
    # -----------------------------
    reports = (
        DailyIncomeReport.objects
        .select_related("member")
        .filter(date=run_date)
        .order_by("member__auto_id")
    )

    # -----------------------------
    # 3) Min/Max total income filter
    # -----------------------------
    min_income = request.GET.get("min")
    max_income = request.GET.get("max")

    if min_income:
        try:
            reports = reports.filter(total_income__gte=Decimal(str(min_income)))
        except Exception:
            pass

    if max_income:
        try:
            reports = reports.filter(total_income__lte=Decimal(str(max_income)))
        except Exception:
            pass

    # -----------------------------
    # 4) Totals (same keys your template expects)
    # -----------------------------
    totals = reports.aggregate(
        total_eligibility=Coalesce(
            Sum("binary_eligibility_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_binary=Coalesce(
            Sum("binary_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_sponsor=Coalesce(
            Sum("sponsor_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_wallet=Coalesce(
            Sum("member__repurchase_wallet"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_salary=Coalesce(
            Sum("salary_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        grand_total=Coalesce(
            Sum("total_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
    )

    # -----------------------------
    # 5) Render (template expects reports/totals/run_date)
    # -----------------------------
    return render(request, "herbalapp/income_report.html", {
        "run_date": run_date,
        "reports": reports,
        "totals": totals,

        # Optional: keep filters to show in UI (if needed)
        "min_income": min_income or "",
        "max_income": max_income or "",
    })

# ======================================================
# EXPORT INCOME TO EXCEL (multi-sheet + charts)
# ======================================================
def export_income_excel(request):
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Border, Side
    from django.http import HttpResponse

    # Load all members
    members = Member.objects.all().order_by("id")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Style presets
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Group by district → taluk → side
    districts = members.values_list("district", flat=True).distinct()

    for district in districts:
        taluks = members.filter(district=district).values_list("taluk", flat=True).distinct()

        for taluk in taluks:
            sides = members.filter(district=district, taluk=taluk).values_list("side", flat=True).distinct()

            for side in sides:

                # Sheet name (max 31 chars)
                sheet_name = f"{district}-{taluk}-{side}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                # Header row
                headers = [
                    "Member ID", "Name", "Joining Package", "Binary Income",
                    "Flash Bonus", "Sponsor Income", "Salary", "Stock Commission",
                    "Total Income", "Rank Title", "Rank Start Date",
                    "Monthly Income", "Duration (Months)", "Months Paid", "Active",
                    "Left BV", "Right BV", "Matched BV",
                    "Repurchase Wallet", "Flash Wallet"
                ]
                ws.append(headers)

                # Style header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.border = border

                # Data rows
                group_members = members.filter(district=district, taluk=taluk, side=side)

                for m in group_members:

                    # Safe income fetch
                    income = getattr(m, 'calculate_full_income', lambda: {})()

                    joining = income.get("joining", 0)
                    binary_income = income.get("binary_income", 0)
                    flash_bonus = income.get("flash_bonus", 0)
                    sponsor_income = income.get("sponsor_income", 0)
                    salary = income.get("salary", 0)
                    stock_commission = income.get("stock_commission", 0)
                    total_income_all = income.get("total_income_all", 0)

                    # BV
                    left_bv, right_bv = m.get_bv_counts()
                    matched_bv = min(left_bv, right_bv)

                    # Wallets
                    repurchase_wallet = getattr(m, "repurchase_wallet", 0)
                    flash_wallet = getattr(m, "flash_wallet", 0)

                    # Rank rewards
                    rewards_qs = RankReward.objects.filter(member=m)

                    if rewards_qs.exists():
                        for r in rewards_qs:
                            row = [
                                m.auto_id, m.name,
                                joining, binary_income, flash_bonus,
                                sponsor_income, salary, stock_commission,
                                total_income_all,
                                r.rank_title, r.start_date, r.monthly_income,
                                r.duration_months, r.months_paid, r.active,
                                left_bv, right_bv, matched_bv,
                                repurchase_wallet, flash_wallet
                            ]
                            ws.append(row)
                    else:
                        row = [
                            m.auto_id, m.name,
                            joining, binary_income, flash_bonus,
                            sponsor_income, salary, stock_commission,
                            total_income_all,
                            None, None, None, None, None, None,
                            left_bv, right_bv, matched_bv,
                            repurchase_wallet, flash_wallet
                        ]
                        ws.append(row)

                # Apply borders to all data rows
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border

                # Rank Summary
                ws.append([])
                ws.append(["Rank Summary"])
                rank_titles = ["1st Star", "Double Star", "Triple Star"]

                for title in rank_titles:
                    count = group_members.filter(current_rank=title).count()
                    ws.append([title, count])

                # BV Summary
                ws.append([])
                ws.append(["BV Summary"])
                total_left = sum(m.get_bv_counts()[0] for m in group_members)
                total_right = sum(m.get_bv_counts()[1] for m in group_members)
                ws.append(["Total Left BV", total_left])
                ws.append(["Total Right BV", total_right])
                ws.append(["Total Matched BV", min(total_left, total_right)])

                # Wallet Summary
                ws.append([])
                ws.append(["Wallet Summary"])
                ws.append(["Total Repurchase Wallet", sum(getattr(m, "repurchase_wallet", 0) for m in group_members)])
                ws.append(["Total Flash Wallet", sum(getattr(m, "flash_wallet", 0) for m in group_members)])

                # Auto column width
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Freeze header
                ws.freeze_panes = "A2"

                # Rank Chart
                chart = BarChart()
                chart.title = "Rank Distribution"
                chart.x_axis.title = "Rank"
                chart.y_axis.title = "Members"

                # Rank summary rows
                rank_start = ws.max_row - (len(rank_titles) + 4)
                rank_end = rank_start + len(rank_titles) - 1

                data_ref = Reference(ws, min_col=2, min_row=rank_start, max_row=rank_end)
                cat_ref = Reference(ws, min_col=1, min_row=rank_start, max_row=rank_end)

                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(cat_ref)
                ws.add_chart(chart, "M2")

    # Dashboard Sheet
    dash = wb.create_sheet("Dashboard")
    dash.append(["Dashboard Summary"])
    dash.append(["Total Members", members.count()])
    dash.append(["Total Income", sum(
        getattr(m, 'calculate_full_income', lambda: {})().get("total_income_all", 0)
        for m in members
    )])

    # Auto width for dashboard
    for col in dash.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        dash.column_dimensions[col_letter].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="income_report_grouped.xlsx"'
    wb.save(response)
    return response

import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from herbalapp.models import Member

def export_members_income(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Income"

    headers = [
        'Member ID', 'Name', 'Joining Package',
        'Binary Income', 'Flash Bonus', 'Sponsor Income',
        'Repurchase Wallet', 'Rank Reward', 'Total Income'
    ]
    ws.append(headers)

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    today = datetime.date.today()
    members = Member.objects.all()

    for member in members:
        if hasattr(member, "income_date") and member.income_date != today:
            binary_income = sponsor_income = flash_bonus = repurchase_wallet = rank_reward = 0
        else:
            binary_income = member.calculate_binary_income()
            sponsor_income = member.sponsor_income
            flash_bonus = member.flash_bonus
            repurchase_wallet = member.repurchase_wallet
            rank_reward = member.calculate_rank_reward()

        total_income = (
            binary_income + sponsor_income + flash_bonus +
            repurchase_wallet + rank_reward
        )

        ws.append([
            member.auto_id,
            member.name,
            member.package,
            binary_income,
            flash_bonus,
            sponsor_income,
            repurchase_wallet,
            rank_reward,
            total_income
        ])

    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = border

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze header
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members_income.xlsx'
    wb.save(response)
    return response

import requests
from django.utils import timezone
from django.shortcuts import render
from .models import Member, DailyIncomeReport

def generate_daily_report():
    today = timezone.now().date()
    reports = []

    for member in Member.objects.all():
        income = member.calculate_full_income()
        report = DailyIncomeReport.objects.create(
            date=today,
            member=member,
            binary_income=income["binary_income"],
            flash_bonus=income["flash_bonus"],
            sponsor_income=income["sponsor_income"],
            salary=income["salary"],
            stock_commission=income["stock_commission"],
            total_income=income["total_income_all"],
        )
        reports.append(report)

    # ✅ Prepare WhatsApp message
    body = f"Rocky Herbals Daily Income Report - {today}\n\n"
    for r in reports:
        body += f"{r.member.auto_id} - {r.member.name} : ₹{r.total_income}\n"

    # ✅ WhatsApp Cloud API details
    WHATSAPP_TOKEN = "YOUR_WHATSAPP_ACCESS_TOKEN"
    PHONE_NUMBER_ID = "YOUR_WHATSAPP_PHONE_NUMBER_ID"
    RECEIVER = "918122105779"  # ✅ Your mobile number with country code

    url = f"https://graph.facebook.com/v17.0/{PHONE_NUMBER_ID}/messages"

    payload = {
        "messaging_product": "whatsapp",
        "to": RECEIVER,
        "type": "text",
        "text": {"body": body}
    }

    headers = {
        "Authorization": f"Bearer {WHATSAPP_TOKEN}",
        "Content-Type": "application/json"
    }

    # ✅ Send WhatsApp message
    requests.post(url, json=payload, headers=headers)

    return "Daily report sent via WhatsApp"

# ✅ Dashboard (Member Home) — FINAL BV + RANK + INCOME SAFE VERSION
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils import timezone
from django.db.models import Sum

from .models import Member, Product, DailyIncomeReport, RankReward


@login_required
def dashboard(request):
    """
    Member dashboard:
    ✅ Shows latest income snapshot (from DailyIncomeReport)
    ✅ Shows BV / Rank progress (from Member totals)
    ✅ Shows rank reward running status (RankReward)
    ✅ Shows products list
    ✅ No template method calls
    """

    # -------------------------------------------------
    # 1) Resolve member for logged in user
    # -------------------------------------------------
    # If you have OneToOne User->Member, this works:
    try:
        member = request.user.member
    except Exception:
        # Fallback: if username equals auto_id (optional)
        member = Member.objects.filter(auto_id=getattr(request.user, "username", "")).first()
        if not member:
            messages.error(request, "Member profile not found.")
            return redirect("member_list")

    # -------------------------------------------------
    # 2) Today's report (or latest report) income snapshot
    # -------------------------------------------------
    today = timezone.localdate()

    report = (
        DailyIncomeReport.objects
        .filter(member=member)
        .order_by("-date")
        .first()
    )

    # Safe defaults
    binary_income = Decimal("0.00")
    eligibility_income = Decimal("0.00")
    sponsor_income = Decimal("0.00")
    flashout_wallet_income = Decimal("0.00")
    salary_income = Decimal("0.00")
    total_income = Decimal("0.00")

    if report:
        binary_income = report.binary_income or Decimal("0.00")
        eligibility_income = report.binary_eligibility_income or Decimal("0.00")
        sponsor_income = report.sponsor_income or Decimal("0.00")
        flashout_wallet_income = report.flashout_wallet_income or Decimal("0.00")
        salary_income = report.salary_income or Decimal("0.00")
        total_income = report.total_income or Decimal("0.00")

    # -------------------------------------------------
    # 3) Wallets (Member table)
    # -------------------------------------------------
    rep_wallet = getattr(member, "repurchase_wallet", Decimal("0.00")) or Decimal("0.00")
    flash_wallet = getattr(member, "flash_wallet", Decimal("0.00")) or Decimal("0.00")
    main_wallet = getattr(member, "main_wallet", Decimal("0.00")) or Decimal("0.00")

    # -------------------------------------------------
    # 4) BV + Matched BV (Member snapshots)
    # -------------------------------------------------
    left_bv = int(getattr(member, "total_left_bv", 0) or 0)
    right_bv = int(getattr(member, "total_right_bv", 0) or 0)
    matched_bv = min(left_bv, right_bv)

    # -------------------------------------------------
    # 5) Rank fields
    # -------------------------------------------------
    rank = (getattr(member, "current_rank", None) or "").strip() or None
    rank_level = int(getattr(member, "rank_level", 0) or 0)
    rank_checkpoint_bv = int(getattr(member, "rank_checkpoint_bv", 0) or 0)

    # Rank Target Logic (lifetime, step-by-step from checkpoint)
    # Level 0 -> First Star target 50k
    # Level 1 -> Double Star target 1L
    # Level 2 -> Triple Star target 2.5L
    if rank_level == 0:
        target_bv = 50000
        next_rank_name = "1st Star"
    elif rank_level == 1:
        target_bv = 100000
        next_rank_name = "Double Star"
    elif rank_level == 2:
        target_bv = 250000
        next_rank_name = "Triple Star"
    else:
        target_bv = 250000
        next_rank_name = "Next Rank"

    # progress is based on (matched_bv - checkpoint)
    progress_current = max(matched_bv - rank_checkpoint_bv, 0)
    target_span = max(target_bv - rank_checkpoint_bv, 1)
    progress_pct = int(min((progress_current * 100) / target_span, 100))

    # -------------------------------------------------
    # 6) RankReward status (if any active)
    # -------------------------------------------------
    active_rewards = RankReward.objects.filter(member=member, active=True).order_by("-start_date")
    reward = active_rewards.first()

    reward_info = None
    if reward:
        reward_info = {
            "rank_title": reward.rank_title,
            "monthly_income": reward.monthly_income,
            "duration_months": reward.duration_months,
            "months_paid": reward.months_paid,
            "active": reward.active,
            "start_date": reward.start_date,
        }

    # -------------------------------------------------
    # 7) Products
    # -------------------------------------------------
    products = Product.objects.all().order_by("-id")[:6]

    context = {
        "member": member,
        "products": products,

        # Income snapshot (latest report)
        "report": report,
        "report_date": report.date if report else today,

        "binary_income": binary_income,
        "eligibility_income": eligibility_income,
        "sponsor_income": sponsor_income,
        "flashout_wallet_income": flashout_wallet_income,
        "salary_income": salary_income,
        "total_income": total_income,

        # Wallet
        "rep_wallet": rep_wallet,
        "flash_wallet": flash_wallet,
        "main_wallet": main_wallet,

        # BV + rank
        "left_bv": left_bv,
        "right_bv": right_bv,
        "matched_bv": matched_bv,
        "rank": rank,
        "rank_level": rank_level,
        "rank_checkpoint_bv": rank_checkpoint_bv,

        "target_bv": target_bv,
        "next_rank_name": next_rank_name,
        "progress_current": progress_current,
        "progress_pct": progress_pct,

        # rank reward info
        "reward_info": reward_info,
    }

    return render(request, "dashboard.html", context)


# ✅ Members list
@login_required
def member_list(request):
    members = Member.objects.all().order_by("auto_id")
    return render(request, "member_list.html", {"members": members})


# ✅ Products page
@login_required
def products(request):
    return render(request, "products.html")


# ✅ Income page
@login_required
def income_page(request):
    return render(request, "income.html")
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from .models import Member


def member_bv(request, auto_id):

    # ✅ Support both numeric PK and business auto_id
    if str(auto_id).isdigit():
        member = get_object_or_404(Member, id=int(auto_id))
    else:
        member = get_object_or_404(Member, auto_id=auto_id)

    # ✅ Calculate BV safely
    bv_data = member.calculate_bv()  # returns dict: self_bv, left_bv, right_bv, total_bv

    return JsonResponse(bv_data)


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Member


def member_register(request):

    if request.method == "POST":

        # -------- BASIC DETAILS --------
        name = request.POST.get("name", "").strip()
        mobile = request.POST.get("mobile", "").strip()
        email = request.POST.get("email", "").strip()
        auto_id = request.POST.get("auto_id", "").strip()

        place = request.POST.get("place", "").strip()
        district = request.POST.get("district", "").strip()
        pincode = request.POST.get("pincode", "").strip()
        aadhar = request.POST.get("aadhar", "").strip()

        # -------- MLM CRITICAL FIELDS --------
        placement_id = request.POST.get("placement_id", "").strip()
        sponsor_id = request.POST.get("sponsor_id", "").strip()
        side = request.POST.get("side", "").strip()   # left / right

        # -------- VALIDATIONS --------
        if not all([name, mobile, auto_id, placement_id, sponsor_id, side]):
            messages.error(request, "All required fields must be filled.")
            return redirect("member_register")

        if side not in ["left", "right"]:
            messages.error(request, "Invalid side selected.")
            return redirect("member_register")

        if len(mobile) != 10 or not mobile.isdigit():
            messages.error(request, "Invalid mobile number.")
            return redirect("member_register")

        if Member.objects.filter(auto_id=auto_id).exists():
            messages.error(request, f"Auto ID {auto_id} already exists.")
            return redirect("member_register")

        # -------- RESOLVE PLACEMENT PARENT --------
        try:
            parent = Member.objects.get(auto_id=placement_id)
        except Member.DoesNotExist:
            messages.error(request, "Invalid placement ID.")
            return redirect("member_register")

        # -------- SAVE MEMBER --------
        Member.objects.create(
            name=name,
            phone=mobile,
            email=email,
            auto_id=auto_id,

            placement_id=placement_id,   # STRING auto_id
            sponsor_id=sponsor_id,       # STRING auto_id
            parent=parent,               # FK
            side=side,                   # left / right

            place=place,
            district=district,
            pincode=pincode,
            aadhar=aadhar,
        )

        messages.success(
            request,
            f"Member {name} added successfully under {placement_id} ({side})."
        )
        return redirect("member_list")

    return render(request, "member_register.html")


# herbalapp/views.py

from django.contrib import messages
from django.db import transaction
from django.utils import timezone

from .models import Member


# ---------------------------------
# AUTO ID GENERATOR (SAFE)
# ---------------------------------
def generate_auto_id():
    last = (
        Member.objects
        .filter(auto_id__startswith="rocky")
        .order_by("-id")
        .first()
    )
    if last and last.auto_id[5:].isdigit():
        return f"rocky{int(last.auto_id[5:]) + 1:03d}"
    return "rocky001"


# ---------------------------------
# ADD MEMBER UNDER PARENT
# ---------------------------------
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from .models import Member


@transaction.atomic
def add_member_under_parent(request, parent_id, position):

    parent = get_object_or_404(Member, auto_id=parent_id)
    position = position.lower()

    if position not in ("left", "right"):
        messages.error(request, "Invalid position.")
        return redirect("member_list")

    # Block duplicate side
    if position == "left" and Member.objects.filter(parent=parent, side="left").exists():
        messages.error(request, f"{parent.auto_id} LEFT side already filled.")
        return redirect("tree_view", parent.auto_id)

    if position == "right" and Member.objects.filter(parent=parent, side="right").exists():
        messages.error(request, f"{parent.auto_id} RIGHT side already filled.")
        return redirect("tree_view", parent.auto_id)

    # -------------------------
    # GET → SHOW FORM
    # -------------------------
    if request.method == "GET":
        return render(request, "herbalapp/add_member_form.html", {
            "parent": parent,
            "parent_auto_id": parent.auto_id,
            "position": position,
            "side": position,
        })

    # -------------------------
    # POST → PROCESS FORM
    # -------------------------
    if request.method == "POST":

        # OTP VALIDATION + DEBUG (SAFE)
        otp = (request.POST.get("otp") or request.POST.get("invite_code") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        sponsor_id = (request.POST.get("sponsor_id") or "").strip()

        print("DEBUG OTP =", otp)
        print("DEBUG phone/sponsor =", phone, sponsor_id)

        ok, msg = validate_and_consume_invite(
            otp,
            phone=phone,
            sponsor_auto_id=sponsor_id,
        )

        print("DEBUG validate =", ok, msg)

        if not ok:
            return render(request, "herbalapp/add_member_form.html", {
                "parent": parent,
                "parent_auto_id": parent.auto_id,
                "position": position,
                "side": position,
                "error": msg,
            })

    # AUTO ID GENERATION (SAFE)
        last_member = Member.objects.select_for_update().order_by("-id").first()
        if last_member and last_member.auto_id:
            last_num = int(last_member.auto_id.replace("rocky", ""))
            auto_id = f"rocky{last_num + 1:03d}"
        else:
            auto_id = "rocky001"

        # POST DATA
        name = (request.POST.get("name") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        email = request.POST.get("email")
        aadhar = request.POST.get("aadhar")
        place = request.POST.get("place")
        district = request.POST.get("district")
        pincode = request.POST.get("pincode")
        sponsor_auto_id = (request.POST.get("sponsor_id") or "").strip()
        avatar = request.FILES.get("avatar")

        if not name:
            return render(request, "herbalapp/add_member_form.html", {
                "parent": parent,
                "parent_auto_id": parent.auto_id,
                "position": position,
                "side": position,
                "error": "Name is required",
            })

        if not phone:
            return render(request, "herbalapp/add_member_form.html", {
                "parent": parent,
                "parent_auto_id": parent.auto_id,
                "position": position,
                "side": position,
                "error": "Phone is required",
            })

        sponsor = Member.objects.filter(auto_id=sponsor_auto_id).first()
        if not sponsor:
            return render(request, "herbalapp/add_member_form.html", {
                "parent": parent,
                "parent_auto_id": parent.auto_id,
                "position": position,
                "side": position,
                "error": "Sponsor not found",
            })

        # CREATE MEMBER
        new_member = Member.objects.create(
            auto_id=auto_id,
            name=name,
            phone=phone,
            email=email,
            aadhar=aadhar,
            place=place,
            district=district,
            pincode=pincode,
            parent=parent,
            placement=parent,
            side=position,
            sponsor=sponsor,
            avatar=avatar,
            joined_date=timezone.localdate(),
            is_active=True,
        )

        messages.success(
            request,
            f"Member {new_member.auto_id} added under {parent.auto_id} ({position.upper()})"
        )

        return redirect("tree_view", parent.auto_id)

# ===================== IMPORTS =====================
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Member
from decimal import Decimal

# ===============================================================
# ADD MEMBER VIEW
# - Auto ID
# - Correct binary placement (parent + side)
# - Sponsor linking
# - Tree-safe (auto side)
# ===============================================================
@transaction.atomic
def add_member_form(request):

    # ===================================================
    # AUTO ID GENERATION (SAFE)
    # ===================================================
    last_member = Member.objects.order_by('-auto_id').first()
    if last_member:
        last_num = int(last_member.auto_id.replace("rocky", ""))
        new_member_id = f"rocky{last_num + 1:03d}"
    else:
        new_member_id = "rocky001"

    # ===================================================
    # PREFILL PARENT FROM TREE CLICK (?parent=rocky005)
    # ===================================================
    parent_code = request.GET.get("parent")
    parent_member = None
    if parent_code:
        parent_member = Member.objects.filter(auto_id=parent_code).first()

    # ===================================================
    # FORM SUBMIT
    # ===================================================
    if request.method == "POST":

        auto_id = request.POST.get("auto_id")
        sponsor_code = request.POST.get("sponsor_id")
        parent_code = request.POST.get("parent_id")  # clicked from tree

        otp = (request.POST.get("otp") or request.POST.get("invite_code") or "").strip()
        phone = request.POST.get("phone", "").strip()  # if your form has phone field
        sponsor_auto_id = request.POST.get("sponsor_id", "").strip()  # if sponsor_id present

        ok, msg = validate_and_consume_invite(
            invite_code,
            phone=phone,
            sponsor_auto_id=sponsor_auto_id
        )

        if not ok:
            messages.error(request, msg)
            return redirect("add_member_form")

        # ---------------- VALIDATIONS ----------------
        if Member.objects.filter(auto_id=auto_id).exists():
            return render(request, "herbalapp/add_member_form.html", {
                "error": "Member ID already exists",
                "auto_id": new_member_id
            })

        sponsor = Member.objects.filter(auto_id=sponsor_code).first()
        parent = Member.objects.filter(auto_id=parent_code).first() or parent_member

        if not parent:
            return render(request, "herbalapp/add_member_form.html", {
                "error": "Invalid parent ID",
                "auto_id": new_member_id
            })

        if not sponsor:
            return render(request, "herbalapp/add_member_form.html", {
                "error": "Invalid sponsor ID",
                "auto_id": new_member_id
            })

        # ===================================================
        # AUTO SIDE & PLACEMENT DETECTION (TREE-SAFE)
        # ===================================================
        def find_next_available(member):
            left = Member.objects.filter(parent=member, side='left').first()
            right = Member.objects.filter(parent=member, side='right').first()

            if not left:
                return member, 'left'
            if not right:
                return member, 'right'

            # Recursively check left subtree
            result = find_next_available(left)
            if result:
                return result
            # Recursively check right subtree
            return find_next_available(right)

        placement, side = find_next_available(parent)

        # =============================
        # SIDE SAFETY (FINAL GUARD)
        # =============================
        if side not in ("left", "right"):
            messages.error(request, "Invalid side detected. Please retry.")
            return redirect("/tree/")


        # ===================================================
        # CREATE MEMBER
        # ===================================================
        new_member = Member.objects.create(
            auto_id=auto_id,
            name=request.POST.get("name"),
            phone=request.POST.get("phone"),
            email=request.POST.get("email"),
            aadhar=request.POST.get("aadhar"),
            place=request.POST.get("place"),
            district=request.POST.get("district"),
            pincode=request.POST.get("pincode"),

            parent=placement,   # 🔥 Binary tree position
            side=side,          # 🔥 left / right
            sponsor=sponsor     # 🔥 Sponsor / referral
        )

        messages.success(
            request,
            f"Member {new_member.auto_id} added under {placement.auto_id} ({side.upper()})"
        )

        return redirect("/tree/")

        # ===================================================
        # INITIAL PAGE LOAD
        # ===================================================
        return render(request, "herbalapp/add_member_form.html", {
            "auto_id": new_member_id,
            "placement_member_id": parent_member.auto_id if parent_member else ""
        })


from django.shortcuts import render, get_object_or_404
from herbalapp.models import Member, DailyIncomeReport
from django.db.models import Sum

def income_chart(request, auto_id):
    # -----------------------------
    # Load member safely
    # -----------------------------
    member = get_object_or_404(Member, auto_id=auto_id)

    # -----------------------------
    # Fetch reports
    # -----------------------------
    reports = (
        DailyIncomeReport.objects
        .filter(member=member)
        .order_by('date')
    )

    # -----------------------------
    # Prepare chart data
    # -----------------------------
    dates = [str(r.date) for r in reports]
    salary = [float(r.salary_income or 0) for r in reports]
    total_income = [float(r.total_income or 0) for r in reports]

    # -----------------------------
    # Optional: cumulative totals
    # -----------------------------
    cumulative_salary = []
    cumulative_total = []
    sum_salary = 0
    sum_total = 0
    for s, t in zip(salary, total_income):
        sum_salary += s
        sum_total += t
        cumulative_salary.append(sum_salary)
        cumulative_total.append(sum_total)

    # -----------------------------
    # Render chart template
    # -----------------------------
    return render(request, "herbalapp/income_chart.html", {
        "member": member,
        "dates": dates,
        "salary": salary,
        "total_income": total_income,
        "cumulative_salary": cumulative_salary,
        "cumulative_total": cumulative_total,
    })

from decimal import Decimal
from datetime import datetime
from django.shortcuts import render
from django.utils.timezone import localdate
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from herbalapp.models import DailyIncomeReport


def income_report(request):
    # -----------------------------
    # Date filter (supports DD-MM-YYYY + YYYY-MM-DD)
    # -----------------------------
    date_str = request.GET.get("date")

    if date_str:
        try:
            # UI sends: 07-02-2026
            run_date = datetime.strptime(date_str, "%d-%m-%Y").date()
        except ValueError:
            try:
                # If browser sends: 2026-02-07
                run_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                run_date = localdate()
    else:
        run_date = localdate()

    # -----------------------------
    # Fetch existing reports ONLY (no creating rows here)
    # -----------------------------
    reports = (
        DailyIncomeReport.objects
        .select_related("member")
        .filter(date=run_date)
        .order_by("member__auto_id")
    )

    # -----------------------------
    # Totals
    # -----------------------------
    totals = reports.aggregate(
        total_eligibility=Coalesce(
            Sum("binary_eligibility_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_binary=Coalesce(
            Sum("binary_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_sponsor=Coalesce(
            Sum("sponsor_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_wallet=Coalesce(
            Sum("member__repurchase_wallet"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_salary=Coalesce(
            Sum("salary_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        grand_total=Coalesce(
            Sum("total_income"),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
    )

    return render(request, "herbalapp/income_report.html", {
        "run_date": run_date,
        "reports": reports,
        "totals": totals,
    })

from django.shortcuts import render, get_object_or_404
from django.db.models import Value
from django.db.models.functions import Coalesce
from .models import Member
from .ranks import get_rank


# ---------------------------------------------------------
# ✅ 1. Rank Report (All Members)
# ---------------------------------------------------------
def rank_report(request):
    # Null-safe BV values + proper ordering
    members = Member.objects.annotate(
        left_bv_safe=Coalesce('total_left_bv', Value(0)),
        right_bv_safe=Coalesce('total_right_bv', Value(0)),
    ).order_by('-left_bv_safe')

    return render(request, "rank_report.html", {
        "members": members
    })


# ---------------------------------------------------------
# ✅ 2. Salary Report (Members with salary > 0)
# ---------------------------------------------------------
def salary_report(request):
    members = Member.objects.filter(salary__gt=0).order_by('-salary', 'name')

    return render(request, "salary_report.html", {
        "members": members
    })


# ---------------------------------------------------------
# ✅ 3. Member Rank Detail Page
# ---------------------------------------------------------
def member_rank_detail(request, auto_id):
    member = get_object_or_404(Member, auto_id=auto_id)

    # Null-safe BV
    left_bv = member.total_left_bv or 0
    right_bv = member.total_right_bv or 0
    matched_bv = min(left_bv, right_bv)

    # Rank engine
    rank_info = get_rank(matched_bv) or {}

    return render(request, "member_rank_detail.html", {
        "member": member,
        "matched_bv": matched_bv,
        "rank_info": rank_info,
    })

from django.http import JsonResponse
from datetime import date
from .models import DailyIncomeReport

@login_required
def income_api(request):
    date_str = request.GET.get("date")

    if date_str:
        report_date = date_str
    else:
        report_date = date.today()

    qs = DailyIncomeReport.objects.filter(date=report_date)

    data = []
    for row in qs:
        data.append({
            "member_id": row.member.auto_id,
            "binary_income": float(row.binary_income),
            "eligibility_income": float(row.binary_eligibility_income),
            "sponsor_income": float(row.sponsor_income),
            "total_income": float(row.total_income),
        })

    return JsonResponse({"data": data})

from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from datetime import timedelta
from herbalapp.models import Member

# ✅ Use your existing engine function (best counts)
from herbalapp.mlm.final_master_engine import count_all_descendants

def tree_modern_page(request, root_id="rocky002"):
    return render(request, "tree.html", {"root_id": root_id})

def tree_data(request, root_id):
    root = get_object_or_404(Member, auto_id=root_id)

    collected = []

    def dfs(parent):
        children = Member.objects.filter(parent=parent, is_active=True).order_by("id")
        for c in children:
            collected.append(c)
            dfs(c)

    dfs(root)

    def avatar_url(m):
        img = getattr(m, "avatar", None)
        if img and hasattr(img, "url"):
            return img.url
        return "/static/img/default-avatar.png"

    nodes = [{
        "data": {
            "id": root.auto_id,
            "name": root.name,
            "bv": str(root.bv or 0),
            "avatar": avatar_url(root),
        }
    }]

    edges = []

    for m in collected:
        nodes.append({
            "data": {
                "id": m.auto_id,
                "name": m.name,
                "bv": str(m.bv or 0),
                "avatar": avatar_url(m),
            }
        })
        if m.parent:
            edges.append({
                "data": {
                    "source": m.parent.auto_id,
                    "target": m.auto_id
                }
            })

    return JsonResponse({"nodes": nodes, "edges": edges})

def member_details_api(request, member_id):
    m = get_object_or_404(Member, auto_id=member_id)

    left_count = count_all_descendants(m, "left", as_of_date=None)
    right_count = count_all_descendants(m, "right", as_of_date=None)

    img = getattr(m, "avatar", None)
    avatar = img.url if img and hasattr(img, "url") else "/static/img/default-avatar.png"

    return JsonResponse({
        "auto_id": m.auto_id,
        "name": m.name,
        "joined_date": str(m.joined_date) if m.joined_date else "",
        "left_count": left_count,
        "right_count": right_count,
        "bv": str(m.bv or 0),
        "avatar": avatar,
        "sponsor": m.sponsor.auto_id if m.sponsor else "",
        "parent": m.parent.auto_id if m.parent else "",
        "side": m.side or "",
    })
from django.utils import timezone
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from herbalapp.models import DailyIncomeReport, RankReward, Member

@login_required
def member_root_income_dashboard(request):
    # ✅ if your auth linked like request.user.member
    try:
        member = request.user.member
    except:
        messages.error(request, "Member profile not found.")
        return redirect("member_list")

    # ✅ BV (life-time)
    bv_data = member.calculate_bv()
    left_bv = bv_data.get("left_bv", 0)
    right_bv = bv_data.get("right_bv", 0)
    matched_bv = bv_data.get("matched_bv", 0)

    # ✅ Lifetime income totals (DailyIncomeReport தான் source of truth)
    totals = DailyIncomeReport.objects.filter(member=member).aggregate(
        total_binary=Coalesce(Sum("binary_income"), Value(0), output_field=DecimalField()),
        total_elig=Coalesce(Sum("binary_eligibility_income"), Value(0), output_field=DecimalField()),
        total_sponsor=Coalesce(Sum("sponsor_income"), Value(0), output_field=DecimalField()),
        total_flash=Coalesce(Sum("flashout_wallet_income"), Value(0), output_field=DecimalField()),
        total_salary=Coalesce(Sum("salary_income"), Value(0), output_field=DecimalField()),
        grand_total=Coalesce(Sum("total_income"), Value(0), output_field=DecimalField()),
    )

    # ✅ Today report
    today = timezone.localdate()
    today_report = DailyIncomeReport.objects.filter(member=member, date=today).first()

    # ✅ Rank reward status (active payouts)
    active_rewards = RankReward.objects.filter(member=member, active=True).order_by("-start_date")
    all_rewards = RankReward.objects.filter(member=member).order_by("-start_date")

    # ✅ Last 30 days history
    last_reports = DailyIncomeReport.objects.filter(member=member).order_by("-date")[:30]

    return render(request, "herbalapp/member_root_income.html", {
        "member": member,
        "wallet": member.main_wallet,

        "left_bv": left_bv,
        "right_bv": right_bv,
        "matched_bv": matched_bv,

        "current_rank": member.current_rank,
        "rank_level": member.rank_level,
        "rank_checkpoint_bv": member.rank_checkpoint_bv,

        "totals": totals,
        "today_report": today_report,
        "active_rewards": active_rewards,
        "all_rewards": all_rewards,
        "last_reports": last_reports,
    })

